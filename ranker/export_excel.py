# ranker/export_excel.py — Excel export + formatting
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from ranker.config import EXPORT_COLS, FRIENDLY_NAMES, PCT_COLS_DECIMAL, PCT_COLS_FRACTION, ALL_PCT_COLS

def style_and_export(df: pd.DataFrame, filepath: str):
    out_df = df.reindex(columns=[c for c in EXPORT_COLS if c in df.columns])
    out_df = out_df.rename(columns=FRIENDLY_NAMES)
    for c in ALL_PCT_COLS:
        fn = FRIENDLY_NAMES.get(c, c)
        if fn in out_df.columns:
            out_df[fn] = (out_df[fn] * 100).round(2)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="Rankings", index=False)
        sector_cols = [c for c in list(PILLAR_MAP.values())
                       + ["composite_score", "valuation_score", "tr_smart_score"]
                       if c in df.columns]
        df.groupby("sector")[sector_cols].agg(["median","mean","count"]).round(1).to_excel(
            writer, sheet_name="Sector Analysis")
        out_df.head(50).to_excel(writer, sheet_name="Top 50", index=False)
        wb = writer.book
        for sn in ["Rankings", "Top 50"]:
            _format_sheet(wb[sn])

    print(f"✅  Excel → {filepath}")


def _format_sheet(ws):
    HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
    HEADER_TR_FILL = PatternFill("solid", fgColor="154360")
    ALT_FILL       = PatternFill("solid", fgColor="EBF3FB")
    BORDER = Border(bottom=Side(style="thin", color="BFBFBF"),
                    right=Side(style="thin",  color="BFBFBF"))

    score_idx = val_idx = smart_idx = None
    for idx, cell in enumerate(ws[1], 1):
        val = str(cell.value or "")
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = BORDER
        cell.fill      = HEADER_TR_FILL if val.startswith("TR") else HEADER_FILL
        if "Composite Score" in val: score_idx = idx
        if "Cheap/Expensive" in val: val_idx   = idx
        if "SmartScore"      in val: smart_idx = idx

    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if ri % 2 == 0:
                cell.fill = ALT_FILL

    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, 28)

    for ci in [score_idx, val_idx, smart_idx]:
        if ci:
            cl = get_column_letter(ci)
            ws.conditional_formatting.add(
                f"{cl}2:{cl}{ws.max_row}",
                ColorScaleRule(start_type="min",       start_color="FF4444",
                               mid_type="percentile",  mid_value=50, mid_color="FFFF00",
                               end_type="max",         end_color="00B050"),
            )
    ws.freeze_panes = "D2"
